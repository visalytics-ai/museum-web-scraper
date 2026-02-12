# Setup (example, macOS/Linux):
#   1) Create a virtual environment in your project folder:
#        python3 -m venv .venv
#   2) Activate it:
#        source .venv/bin/activate
#   3) Install dependencies (see requirements.txt / pyproject.toml):
#        pip install -r requirements.txt
#
# Notes:
# - Avoid hard-coding your personal file paths/usernames in this script.
# - You can override common paths via environment variables (see CONFIGURATION).

import os
import re
import time
from typing import Dict, List, Any, Optional

import requests
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright, Page
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------------

API_BASE_URL = "https://collectionapi.metmuseum.org/public/collection/v1/objects/{}"
SEARCH_URL = "https://collectionapi.metmuseum.org/public/collection/v1/search"

MAX_ADDITIONAL_IMAGES = 8               # store up to this many additional image URLs/paths
IMAGE_ROOT_DIR = os.environ.get("IMAGE_ROOT_DIR", "downloaded_images")  # root directory where images are saved

TAB_NAMES = [
    "Overview",
    "Signatures, Inscriptions, and Markings",
    "Provenance",
    "References",
]

NON_IMAGE_FIELDS = [
    "objectID",
    "objectName",
    "title",
    "objectBeginDate",
    "objectEndDate",
    "objectDate",
    "culture",
    "period",
    "dynasty",
    "reign",
    "artistDisplayName",
    "artistDisplayBio",
    "medium",
    "dimensions",
    "classification",
    "department",
    "creditLine",
    "repository",
    "objectURL",
    "longDescription",
    "artworkOverviewText",
    "signaturesInscriptionsMarkingsText",
    "provenanceText",
    "referencesText",
]

# Excel-illegal character regex: 0x00–0x08, 0x0B–0x0C, 0x0E–0x1F
ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")

def clean_for_excel(value):
    """Remove Excel-illegal control characters; leave other types unchanged."""
    if isinstance(value, str):
        return ILLEGAL_CHARS_RE.sub("", value)
    return value


# ---------------------------------------------------------------------------
# SEARCH: get object IDs
# ---------------------------------------------------------------------------

def fetch_object_ids_from_search(limit: Optional[int] = None) -> List[int]:
    params = {
        "hasImages": "true",
        "departmentId": 4,
        "q": "sword",
    }
    resp = requests.get(SEARCH_URL, params=params, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    all_ids = data.get("objectIDs") or []
    if limit is not None:
        return all_ids[:limit]
    return all_ids


# ---------------------------------------------------------------------------
# LONG DESCRIPTION SCRAPER (from HTML, robust)
# ---------------------------------------------------------------------------

def fetch_long_description_from_html(object_url: str, timeout: int = 20) -> str:
    if not object_url:
        return ""

    try:
        resp = requests.get(object_url, timeout=timeout)
        if resp.status_code != 200:
            return ""
    except Exception:
        return ""

    soup = BeautifulSoup(resp.text, "html.parser")
    long_text = None
    spans = soup.select('span[data-sentry-component="LegacyOrMarkdownParser"]')

    # Preferred: spans under read-more-wrapper
    for span in spans:
        ancestor = span
        for _ in range(6):
            ancestor = ancestor.parent
            if not ancestor:
                break
            classes = ancestor.get("class") or []
            if any("read-more-wrapper" in c for c in classes):
                text = span.get_text(" ", strip=True)
                if text:
                    long_text = text
                    break
        if long_text:
            break

    # Fallback: longest span with >30 words
    if not long_text:
        best = ""
        best_len = 0
        for span in spans:
            t = span.get_text(" ", strip=True)
            if t:
                wc = len(t.split())
                if wc > best_len and wc > 30:
                    best = t
                    best_len = wc
        if best:
            long_text = best

    # Meta description fallback
    if not long_text:
        meta = soup.find("meta", {"name": "description"})
        if meta and meta.get("content"):
            long_text = meta["content"].strip()

    # OG description fallback
    if not long_text:
        og = soup.find("meta", {"property": "og:description"})
        if og and og.get("content"):
            long_text = og["content"].strip()

    # First 3 paragraphs fallback
    if not long_text:
        paras = soup.find_all("p")
        texts = [p.get_text(" ", strip=True) for p in paras if p.get_text(strip=True)]
        if texts:
            long_text = "\n\n".join(texts[:3])

    return long_text or ""


# ---------------------------------------------------------------------------
# PLAYWRIGHT HELPERS (tabs, title, images)
# ---------------------------------------------------------------------------

def get_artwork_details_section_text(page: Page) -> str:
    text = page.evaluate(
        """
() => {
  const headings = Array.from(document.querySelectorAll('h1, h2, h3'));
  const heading = headings.find(h => h.textContent.includes('Artwork Details'));
  if (!heading) return '';
  const container = heading.closest('section, div') || heading.parentElement;
  return container ? container.innerText : '';
}
"""
    )
    return (text or "").strip()

def clean_panel_text(raw: str) -> str:
    if not raw:
        return ""
    lines = [ln.strip() for ln in raw.splitlines()]
    cleaned = [
        ln for ln in lines
        if ln and ln not in ["Artwork Details", "Object Information"] and ln not in TAB_NAMES
    ]
    return "\n".join(cleaned).strip()

def get_page_title(page: Page) -> str:
    title = page.evaluate(
        """
() => {
  const main = document.querySelector('main') || document.body;
  const h1 = main.querySelector('h1');
  return h1 ? h1.textContent.trim() : (document.title || '');
}
"""
    )
    return (title or "").strip()

def get_page_images(page: Page) -> List[str]:
    urls = page.evaluate(
        """
() => {
  const imgs = Array.from(document.querySelectorAll('main img, body img'));
  const urls = imgs.map(img => img.currentSrc || img.src || '').filter(u => u.startsWith('http'));
  return Array.from(new Set(urls));
}
"""
    )
    return urls or []


def scrape_tabs_for_object(object_id: int, page: Page) -> Dict[str, Any]:
    url = f"https://www.metmuseum.org/art/collection/search/{object_id}"

    try:
        page.goto(url, wait_until="domcontentloaded", timeout=90000)
        page.wait_for_timeout(1500)
    except Exception as e:
        print(f"[WARN] Failed to load {object_id}: {e}")
        return {
            "pageTitle": "",
            "pageImageURLs": [],
            "artworkOverviewText": "",
            "signaturesInscriptionsMarkingsText": "",
            "provenanceText": "",
            "referencesText": "",
            "objectURL_page": url,
        }

    result = {
        "pageTitle": get_page_title(page),
        "pageImageURLs": get_page_images(page),
        "artworkOverviewText": "",
        "signaturesInscriptionsMarkingsText": "",
        "provenanceText": "",
        "referencesText": "",
        "objectURL_page": url,
    }

    for tab_label in TAB_NAMES:
        loc = page.get_by_text(tab_label, exact=True)
        try:
            if loc.count() == 0:
                continue
            loc.first.click(force=True, timeout=5000)
        except Exception:
            print(f"[WARN] Tab '{tab_label}' failed for {object_id}")
            continue

        page.wait_for_timeout(800)
        raw = get_artwork_details_section_text(page)
        cleaned = clean_panel_text(raw)

        if tab_label == "Overview":
            result["artworkOverviewText"] = cleaned
        elif tab_label == "Signatures, Inscriptions, and Markings":
            result["signaturesInscriptionsMarkingsText"] = cleaned
        elif tab_label == "Provenance":
            result["provenanceText"] = cleaned
        elif tab_label == "References":
            result["referencesText"] = cleaned

    # remove duplicate-tab clones
    if result["signaturesInscriptionsMarkingsText"] == result["artworkOverviewText"]:
        result["signaturesInscriptionsMarkingsText"] = ""
    if result["referencesText"] == result["provenanceText"]:
        result["referencesText"] = ""

    return result


# ---------------------------------------------------------------------------
# API
# ---------------------------------------------------------------------------

def fetch_api_data(object_id: int) -> Dict[str, Any]:
    try:
        resp = requests.get(API_BASE_URL.format(object_id), timeout=20)
    except Exception:
        return {}
    if resp.status_code != 200:
        return {}
    try:
        return resp.json()
    except Exception:
        return {}


# ---------------------------------------------------------------------------
# XLSX HELPERS
# ---------------------------------------------------------------------------

def parse_ext(url: str) -> str:
    ext = url.split("?")[0].split(".")[-1].lower()
    return ext if ext in ("jpg", "jpeg", "png", "gif", "tif", "tiff", "bmp") else "jpg"

def embed_image(ws, row, col, url, filepath, w, h):
    """Download and embed a single image (primary thumbnail)."""
    if not url or not filepath or not col:
        return
    try:
        if not os.path.exists(filepath):
            r = requests.get(url, timeout=20)
            if r.status_code != 200:
                return
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            with open(filepath, "wb") as f:
                f.write(r.content)
        img = XLImage(filepath)
        img.width = w
        img.height = h
        ws.add_image(img, f"{col}{row}")
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, h * 0.8)
    except Exception:
        return

def write_headers(ws):
    add_url_cols = [f"additionalImage_{i+1}_URL" for i in range(MAX_ADDITIONAL_IMAGES)]
    add_path_cols = [f"additionalImage_{i+1}_LocalPath" for i in range(MAX_ADDITIONAL_IMAGES)]

    headers = (
        ["primaryImageThumbnail"]
        + NON_IMAGE_FIELDS
        + ["primaryImageURL", "primaryImageLocalPath"]
        + add_url_cols
        + add_path_cols
    )

    ws.append(headers)

    primary_col = None

    for idx, h in enumerate(headers, start=1):
        col = get_column_letter(idx)
        if h == "primaryImageThumbnail":
            primary_col = col

    # Basic widths for readability
    if primary_col:
        ws.column_dimensions[primary_col].width = 20

    if "primaryImageURL" in headers:
        col = get_column_letter(headers.index("primaryImageURL") + 1)
        ws.column_dimensions[col].width = 60

    if "primaryImageLocalPath" in headers:
        col = get_column_letter(headers.index("primaryImageLocalPath") + 1)
        ws.column_dimensions[col].width = 60

    return headers, primary_col


# ---------------------------------------------------------------------------
# MAIN SCRAPER
# ---------------------------------------------------------------------------

def full_scrape(
    output_xlsx: str,
    save_every: int = 25,
    limit: Optional[int] = None,
    start_offset: int = 0,
):
    """
    start_offset: number of IDs to skip from the beginning of the search result.
      e.g., if the script died after 90 objects, set start_offset=90 to start
      at the 91st object ID.
    """
    all_ids = fetch_object_ids_from_search(limit)
    total_all = len(all_ids)

    if start_offset < 0:
        start_offset = 0
    if start_offset > total_all:
        start_offset = total_all

    object_ids = all_ids[start_offset:]
    total = len(object_ids)

    print(f"Total IDs returned by search: {total_all}")
    print(f"Starting from offset {start_offset}, scraping {total} objects.")

    os.makedirs(IMAGE_ROOT_DIR, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Met Swords"

    headers, primary_thumb_col = write_headers(ws)

    row_idx = 2
    processed = 0

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        for idx, object_id in enumerate(object_ids, start=1):
            # idx here is "relative" to start_offset
            absolute_index = start_offset + idx
            print(f"[{absolute_index}/{total_all}] Scraping {object_id}...", flush=True)

            api_data = fetch_api_data(object_id)

            # Tab scraping with fallback
            try:
                tab_data = scrape_tabs_for_object(object_id, page)
            except Exception as e:
                print(f"[ERROR] scrape_tabs_for_object failed for {object_id}: {e}")
                tab_data = {
                    "pageTitle": "",
                    "pageImageURLs": [],
                    "artworkOverviewText": "",
                    "signaturesInscriptionsMarkingsText": "",
                    "provenanceText": "",
                    "referencesText": "",
                    "objectURL_page": f"https://www.metmuseum.org/art/collection/search/{object_id}",
                }

            # Long description
            object_url_html = (
                api_data.get("objectURL")
                if api_data and api_data.get("objectURL")
                else tab_data["objectURL_page"]
            )
            long_desc = fetch_long_description_from_html(object_url_html)

            record: Dict[str, Any] = {}

            # ============ IMAGE URLs ============
            api_primary = api_data.get("primaryImage", "") if api_data else ""
            api_additional = api_data.get("additionalImages") if api_data else []

            if not isinstance(api_additional, list):
                api_additional = list(api_additional)

            page_images = tab_data.get("pageImageURLs", []) or []

            primary_image = api_primary or (page_images[0] if page_images else "")
            additional_images = api_additional or page_images[1:]
            additional_images = additional_images[:MAX_ADDITIONAL_IMAGES]

            record["primaryImage"] = primary_image
            record["additionalImages"] = additional_images

            # ============ NON-IMAGE FIELDS ============
            for field in NON_IMAGE_FIELDS:
                if field == "objectID":
                    record[field] = api_data.get("objectID", object_id)
                elif field == "title":
                    record[field] = api_data.get("title", "") or tab_data.get("pageTitle", "")
                elif field == "objectURL":
                    record[field] = api_data.get("objectURL", "") or object_url_html
                elif field == "longDescription":
                    record[field] = long_desc
                elif field in [
                    "artworkOverviewText",
                    "signaturesInscriptionsMarkingsText",
                    "provenanceText",
                    "referencesText",
                ]:
                    record[field] = tab_data.get(field, "")
                else:
                    record[field] = api_data.get(field, "") if api_data else ""

            # ============ LOCAL IMAGE PATHS ============
            oid_str = str(record["objectID"])
            object_folder = os.path.join(IMAGE_ROOT_DIR, oid_str)
            os.makedirs(object_folder, exist_ok=True)

            # primary
            if primary_image:
                ext = parse_ext(primary_image)
                primary_local = os.path.join(IMAGE_ROOT_DIR, oid_str, f"{oid_str}_1.{ext}")
                primary_local_abs = os.path.join(object_folder, f"{oid_str}_1.{ext}")
            else:
                primary_local = ""
                primary_local_abs = ""

            record["primaryImageLocalPath"] = primary_local

            # additional
            add_local_paths = []
            add_local_abs = []
            for offset, url in enumerate(additional_images, start=2):
                ext = parse_ext(url)
                fname = f"{oid_str}_{offset}.{ext}"
                relp = os.path.join(IMAGE_ROOT_DIR, oid_str, fname)
                absp = os.path.join(object_folder, fname)
                add_local_paths.append(relp)
                add_local_abs.append(absp)

            record["additionalImagesLocalPaths"] = add_local_paths

            # ============ WRITE ROW ============
            row: List[Any] = []

            # primary thumb placeholder (actual image added after append)
            row.append("")

            # non-image columns
            for field in NON_IMAGE_FIELDS:
                row.append(record.get(field, ""))

            # primary URL + local path
            row.append(record.get("primaryImage", ""))
            row.append(record.get("primaryImageLocalPath", ""))

            # additional URLs
            for i in range(MAX_ADDITIONAL_IMAGES):
                row.append(additional_images[i] if i < len(additional_images) else "")

            # additional local paths
            for i in range(MAX_ADDITIONAL_IMAGES):
                row.append(add_local_paths[i] if i < len(add_local_paths) else "")

            # Clean row values for Excel
            row = [clean_for_excel(v) for v in row]

            # Append row
            ws.append(row)

            # ============ EMBED PRIMARY THUMBNAIL ONLY ============
            if primary_image and primary_local_abs:
                embed_image(ws, row_idx, primary_thumb_col, primary_image, primary_local_abs, 150, 150)

            row_idx += 1
            processed += 1

            # Periodic save
            if processed % save_every == 0:
                print(f"Saving checkpoint at {processed} objects...")
                wb.save(output_xlsx)
                print("Checkpoint saved.")

            time.sleep(0.5)

        browser.close()

    # Final save
    print("Final save...")
    wb.save(output_xlsx)
    print("Done.")


# ---------------------------------------------------------------------------
# ENTRY POINT
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    # Runtime configuration (optional):
    #   IMAGE_ROOT_DIR: where downloaded images will be stored (default: ./downloaded_images)
    #   OUTPUT_XLSX:    output Excel filename or path (default: met_swords_full.xlsx)
    #   START_OFFSET:   number of initial IDs to skip (default: 0)
    #
    # Examples:
    #   OUTPUT_XLSX="met_swords_test.xlsx" START_OFFSET=0 python met_swords_full_scrape.py
    #   IMAGE_ROOT_DIR="/path/to/data/images" OUTPUT_XLSX="/path/to/out.xlsx" python met_swords_full_scrape.py

    output_xlsx = os.environ.get("OUTPUT_XLSX", "met_swords_full.xlsx")
    try:
        start_offset = int(os.environ.get("START_OFFSET", "0"))
    except ValueError:
        start_offset = 0

    # Full run (default):
    full_scrape(output_xlsx, save_every=25, limit=None, start_offset=start_offset)

    # Small test (uncomment to use):
    # full_scrape("met_swords_test.xlsx", save_every=5, limit=50, start_offset=0)
